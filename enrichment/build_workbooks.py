#!/usr/bin/env python3
"""Build polished Excel workbooks for all enriched cohorts + a funding-tiered
company prioritization sheet. Adds a funding_tier column to every person row.

Tiers (by capital raised, classified using company_info to separate
investors / public / government from operating-company raises):
  Tier 0  substantial   >= $500M
  Tier 1  a lot         $100M-$500M
  Tier 2  some          < $100M (real disclosed round)
  Tier 3  none/no proof  unknown / bootstrapped
  Public / Investor / Established / Government / Non-profit  (not venture-funded)
"""
import csv, re, unicodedata
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = "/Users/shaurya/raise-summit-2026-data/enrichment"
FILES = [("registrants", "raise_enrichment_full.csv"), ("speakers", "raise_speakers_enriched.csv"),
         ("cxo", "raise_cxo_enriched.csv")]

# ---------- funding parsing / classification ----------
SCALE = {'b':1e9,'bn':1e9,'billion':1e9,'m':1e6,'mm':1e6,'million':1e6,'k':1e3,'thousand':1e3}
CUR = re.compile(r'(?:\$|€|£|aed|usd|eur|gbp)\s*([\d]+(?:[.,]\d+)?)\s*(billion|bn|b|million|mm|m|k|thousand)?', re.I)
WORD = re.compile(r'([\d]+(?:[.,]\d+)?)\s*(billion|million)', re.I)
INVESTOR = ("venture capital","private equity","asset manager","asset management","investment firm",
            "investment management","sovereign wealth","hedge fund","family office","fund manager",
            "portfolio companies","invests in","vc firm","pe firm","buyout","capital partners","wealth fund","pension fund")
PUBLIC = ("publicly traded","publicly-traded","listed on","nyse","nasdaq","stock exchange","public company",
          "trades under","ticker","euronext","ftse","is a public")
GOV = ("government","ministry","state-owned","public sector","statutory body","municipal","national agency")
NONPROF = ("non-profit","nonprofit","ngo","humanitarian","charity")
JUNK = {"confidential","freelance","self-employed","self employed","independent","none","n/a","na","retired",
        "student","unemployed","private","stealth","-","various","/","independent consultant","tbd"}
CEIL = 30e9

def amount(*texts):
    best = 0.0
    for t in texts:
        if not t: continue
        for m in CUR.finditer(t):
            try: v = float(m.group(1).replace(',',''))
            except: continue
            best = max(best, v*SCALE.get((m.group(2) or '').lower(),1))
        for m in WORD.finditer(t):
            try: v = float(m.group(1).replace(',',''))
            except: continue
            best = max(best, v*SCALE.get(m.group(2).lower(),1))
    return best

def classify(total, latest, info):
    t = (total or '').lower(); ci = (info or '').lower()
    if 'public company' in t or any(k in ci for k in PUBLIC): return 'Public', 0
    if any(k in ci for k in INVESTOR): return 'Investor', amount(total, latest)
    if 'government' in t or any(k in ci for k in GOV): return 'Government', 0
    if 'non-profit' in t or 'nonprofit' in t or any(k in ci for k in NONPROF): return 'Non-profit', 0
    amt = amount(total, latest)
    if amt > CEIL: return 'Established', amt
    if amt >= 500e6: return 'Tier 0', amt
    if amt >= 100e6: return 'Tier 1', amt
    if amt > 0: return 'Tier 2', amt
    return 'Tier 3', 0

def norm(s):
    s = unicodedata.normalize('NFKD', s or '').encode('ascii','ignore').decode()
    return re.sub(r'[^a-z0-9 ]','', s.lower()).strip()

# ---------- build company tier map ----------
comp = defaultdict(lambda: {"name":"","total":"","latest":"","inv":"","info":"","home":"","people":0,"cohorts":set(),"amt":0.0})
data = {}
for cohort, f in FILES:
    rows = list(csv.DictReader(open(f"{ROOT}/{f}", encoding="utf-8")))
    data[cohort] = rows
    for r in rows:
        cn = r.get('company_name','').strip()
        if not cn or norm(cn) in JUNK: continue
        c = comp[norm(cn)]
        c["people"] += 1; c["cohorts"].add(cohort)
        if not c["name"]: c["name"] = cn
        if not c["info"] and r.get('company_info','').strip(): c["info"] = r['company_info']
        a = amount(r.get('total_funding',''), r.get('latest_round',''))
        if a > c["amt"] or not c["total"]:
            c["amt"] = max(a, c["amt"]); c["total"] = r.get('total_funding',''); c["latest"] = r.get('latest_round',''); c["inv"] = r.get('key_investors','')
        if not c["home"] and r.get('homepage_url','').strip(): c["home"] = r['homepage_url']

tier_map = {}
for k, c in comp.items():
    tier, amt = classify(c["total"], c["latest"], c["info"])
    tier_map[k] = (tier, amt)

# ---------- styling helpers ----------
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BAND = PatternFill("solid", fgColor="F2F5FA")
LINK_FONT = Font(color="1155CC", underline="single")
TIER_FILL = {
    "Tier 0": "1E7145", "Tier 1": "4CAF50", "Tier 2": "C6E0B4", "Tier 3": "F8CBAD",
    "Public": "9DC3E6", "Investor": "B4A7D6", "Established": "D9D9D9",
    "Government": "FFE599", "Non-profit": "FFD966",
}
TIER_WHITE = {"Tier 0", "Tier 1", "Investor"}
THIN = Side(style="thin", color="DDDDDD")

def style(ws, headers, rows, link_cols=(), wrap_cols=(), tier_col=None, widths=None, header_color="1F3864"):
    ws.append(headers)
    for r in rows: ws.append(r)
    n = len(rows)
    hf = PatternFill("solid", fgColor=header_color)
    for ci, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci); c.fill = hf; c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
        c.border = Border(bottom=Side(style="medium", color="FFFFFF"))
    ws.row_dimensions[1].height = 26
    for ci, h in enumerate(headers, 1):
        L = get_column_letter(ci)
        ws.column_dimensions[L].width = (widths or {}).get(h, 18)
    for ri in range(2, n+2):
        band = (ri % 2 == 0)
        for ci in range(1, len(headers)+1):
            cell = ws.cell(row=ri, column=ci)
            if band: cell.fill = BAND
            cell.border = Border(bottom=THIN)
            h = headers[ci-1]
            if h in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="center")
            if h in link_cols and isinstance(cell.value, str) and cell.value.startswith("http"):
                cell.hyperlink = cell.value; cell.font = LINK_FONT
        if tier_col:
            tc = headers.index(tier_col) + 1
            cell = ws.cell(row=ri, column=tc); v = (cell.value or "")
            if v in TIER_FILL:
                cell.fill = PatternFill("solid", fgColor=TIER_FILL[v])
                cell.font = Font(bold=True, color="FFFFFF" if v in TIER_WHITE else "000000")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n+1}"

# ---------- people workbooks (add funding_tier) ----------
def people_wb(cohort, fname, sheet, header_color):
    rows = data[cohort]
    out = []
    for r in rows:
        tier, amt = tier_map.get(norm(r.get('company_name','')), ("Tier 3", 0))
        out.append(r | {"funding_tier": tier})
    cols = ["first_name","last_name","company_title","company_name","funding_tier","linkedin_url",
            "person_info","homepage_url","company_info","company_goals","total_funding","latest_round","key_investors"]
    head = {"first_name":"First","last_name":"Last","company_title":"Title","company_name":"Company",
            "funding_tier":"Funding Tier","linkedin_url":"LinkedIn","person_info":"Person Info","homepage_url":"Homepage",
            "company_info":"Company Info","company_goals":"Company Goals","total_funding":"Total Funding",
            "latest_round":"Latest Round","key_investors":"Key Investors"}
    cols = [c for c in cols if c in (set(out[0].keys()) | {"funding_tier"})]
    headers = [head[c] for c in cols]
    table = [[r.get(c,"") for c in cols] for r in out]
    widths = {"First":13,"Last":15,"Title":26,"Company":22,"Funding Tier":13,"LinkedIn":40,"Person Info":50,
              "Homepage":34,"Company Info":54,"Company Goals":56,"Total Funding":18,"Latest Round":24,"Key Investors":30}
    wb = Workbook(); style(wb.active, headers, table, link_cols={"LinkedIn","Homepage"},
        wrap_cols={"Person Info","Company Info","Company Goals","Key Investors","Latest Round"},
        tier_col="Funding Tier", widths=widths, header_color=header_color)
    wb.active.title = sheet
    # also write csv with tier
    with open(f"{ROOT}/{fname}", "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=cols); w.writeheader()
        for r in out: w.writerow({c: r.get(c,"") for c in cols})
    wb.save(f"{ROOT}/{fname.replace('.csv','.xlsx')}")
    print(f"{sheet}: {len(table)} rows")

people_wb("registrants", "raise_enrichment_full.csv", "Registrants (4403)", "1F3864")
people_wb("speakers", "raise_speakers_enriched.csv", "Speakers (372)", "6B2C91")
people_wb("cxo", "raise_cxo_enriched.csv", "CXO Summit (260)", "8A1538")

# ---------- companies_by_funding workbook ----------
crows = []
for k, c in comp.items():
    tier, amt = tier_map[k]
    crows.append({"company":c["name"],"tier":tier,"amount_usd":int(amt) if amt else "",
                  "total_funding":c["total"],"latest_round":c["latest"],"key_investors":c["inv"],
                  "people":c["people"],"cohorts":",".join(sorted(c["cohorts"])),
                  "company_info":c["info"],"homepage":c["home"]})
order = {"Tier 0":0,"Tier 1":1,"Tier 2":2,"Public":3,"Investor":4,"Established":5,"Government":6,"Non-profit":7,"Tier 3":8}
crows.sort(key=lambda r:(order[r["tier"]], -(r["amount_usd"] or 0), r["company"].lower()))
cols = ["company","tier","amount_usd","total_funding","latest_round","key_investors","people","cohorts","company_info","homepage"]
with open(f"{ROOT}/companies_by_funding.csv","w",newline="",encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=cols); w.writeheader(); w.writerows(crows)
HEAD = {"company":"Company","tier":"Tier","amount_usd":"Funding (USD)","total_funding":"Total Funding",
        "latest_round":"Latest Round","key_investors":"Key Investors","people":"# People","cohorts":"Cohorts",
        "company_info":"Company Info","homepage":"Homepage"}
headers = [HEAD[c] for c in cols]
table = [[r.get(c,"") for c in cols] for r in crows]
widths = {"Company":28,"Tier":12,"Funding (USD)":16,"Total Funding":34,"Latest Round":24,"Key Investors":30,
          "# People":9,"Cohorts":18,"Company Info":56,"Homepage":34}
wb = Workbook(); ws = wb.active; ws.title = "Companies by Funding"
style(ws, headers, table, link_cols={"Homepage"}, wrap_cols={"Total Funding","Latest Round","Key Investors","Company Info"},
      tier_col="Tier", widths=widths, header_color="0B5394")
# format funding amount as $ millions/billions
amt_col = headers.index("Funding (USD)") + 1
for ri in range(2, len(table)+2):
    cell = ws.cell(row=ri, column=amt_col)
    if isinstance(cell.value, int): cell.number_format = '$#,##0'
wb.save(f"{ROOT}/companies_by_funding.xlsx")
print("Companies by Funding:", len(crows), "| dist:", dict(Counter(r["tier"] for r in crows)))
