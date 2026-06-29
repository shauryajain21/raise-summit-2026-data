#!/usr/bin/env python3
"""Build formatted Excel workbooks from the scraped CSVs.

Produces, in excel/:
  - one workbook per dataset (raise_<name>.xlsx)
  - a combined master workbook (RAISE_Summit_2026_MASTER.xlsx) with one tab each
Formatting: bold/colored header, frozen header row, autofilter, tuned column
widths, wrapped long-text columns.
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = os.path.join(os.path.dirname(__file__), "data")
OUT = os.path.join(os.path.dirname(__file__), "excel")
os.makedirs(OUT, exist_ok=True)

DATASETS = [
    ("registrants", "raise_registrants.csv", "All Registrants"),
    ("attendees", "raise_attendees.csv", "Networking Attendees"),
    ("sponsors", "raise_sponsors.csv", "Sponsors"),
    ("speakers", "raise_speakers.csv", "Speakers"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
WRAP_COLS = {"bio", "pitch", "subtitle"}
MAX_WIDTH = 60
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(bottom=THIN)


def style_sheet(ws, headers, rows):
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])
    # header styling
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci)
        c.fill, c.font = HEADER_FILL, HEADER_FONT
        c.alignment = Alignment(vertical="center", horizontal="left")
    ws.row_dimensions[1].height = 22
    # column widths + wrap
    for ci, h in enumerate(headers, 1):
        letter = get_column_letter(ci)
        if h in WRAP_COLS:
            ws.column_dimensions[letter].width = MAX_WIDTH
            for ri in range(2, len(rows) + 2):
                ws.cell(row=ri, column=ci).alignment = Alignment(wrap_text=True, vertical="top")
        else:
            longest = max([len(h)] + [len(str(r.get(h, ""))) for r in rows[:500]])
            ws.column_dimensions[letter].width = min(max(longest + 2, 10), 45)
    ws.freeze_panes = "A2"
    if rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows)+1}"


def load(csv_path):
    with open(csv_path, encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    headers = list(rows[0].keys()) if rows else []
    return headers, rows


def main():
    master = Workbook()
    master.remove(master.active)
    for key, fname, sheet_title in DATASETS:
        headers, rows = load(os.path.join(DATA, fname))
        # individual workbook
        wb = Workbook()
        style_sheet(wb.active, headers, rows)
        wb.active.title = sheet_title[:31]
        out_path = os.path.join(OUT, fname.replace(".csv", ".xlsx"))
        wb.save(out_path)
        # master tab
        style_sheet(master.create_sheet(sheet_title[:31]), headers, rows)
        print(f"{sheet_title}: {len(rows)} rows -> {os.path.basename(out_path)}")
    master.save(os.path.join(OUT, "RAISE_Summit_2026_MASTER.xlsx"))
    print("master workbook -> RAISE_Summit_2026_MASTER.xlsx")


if __name__ == "__main__":
    main()
